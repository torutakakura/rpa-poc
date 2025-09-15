#!/usr/bin/env python3
"""
RPA Agent - JSON-RPC over stdio server
Electronからの要求を受けて、Python側でRPA処理を実行
"""

import asyncio
import json
import sys
import threading
import time
import traceback
import uuid
from dataclasses import asdict, dataclass
from typing import Any, Dict, Optional

from operation_manager import OperationManager

# Excel操作用（オプション）
try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@dataclass
class JsonRpcRequest:
    """JSON-RPC 2.0 リクエスト"""

    jsonrpc: str = "2.0"
    method: str = ""
    params: Dict[str, Any] = None
    id: Optional[Any] = None


@dataclass
class JsonRpcResponse:
    """JSON-RPC 2.0 レスポンス"""

    jsonrpc: str = "2.0"
    result: Any = None
    error: Optional[Dict[str, Any]] = None
    id: Optional[Any] = None


@dataclass
class JsonRpcNotification:
    """JSON-RPC 2.0 通知（レスポンス不要）"""

    jsonrpc: str = "2.0"
    method: str = ""
    params: Dict[str, Any] = None


class RPAAgent:
    """RPA処理を実行するエージェント"""

    def __init__(self):
        """初期化"""
        self.running = True
        self.operation_manager = OperationManager()
        self.current_task_id = None
        self.task_status = {}

    def start(self):
        """エージェントを開始"""
        # 初期化成功を通知
        self.send_notification("agent.ready", {"status": "ready"})

        # メインループ
        while self.running:
            try:
                # 標準入力から1行読み込み
                line = sys.stdin.readline()
                if not line:
                    break

                # JSON-RPCリクエストをパース
                try:
                    data = json.loads(line.strip())
                    request = JsonRpcRequest(**data)
                except json.JSONDecodeError as e:
                    self.send_error_response(
                        None, -32700, f"Parse error: {str(e)}"
                    )
                    continue
                except Exception as e:
                    self.send_error_response(
                        None, -32600, f"Invalid Request: {str(e)}"
                    )
                    continue

                # リクエストを処理
                threading.Thread(
                    target=self.handle_request, args=(request,), daemon=True
                ).start()

            except KeyboardInterrupt:
                break
            except Exception as e:
                self.send_notification(
                    "agent.error", {"error": str(e), "traceback": traceback.format_exc()}
                )

    def handle_request(self, request: JsonRpcRequest):
        """リクエストを処理"""
        try:
            # メソッドに応じて処理を振り分け
            if request.method == "ping":
                self.handle_ping(request)
            elif request.method == "execute":
                self.handle_execute(request)
            elif request.method == "executeExcel":
                self.handle_execute_excel(request)
            elif request.method == "listOperations":
                self.handle_list_operations(request)
            elif request.method == "getOperationTemplate":
                self.handle_get_operation_template(request)
            elif request.method == "cancelTask":
                self.handle_cancel_task(request)
            elif request.method == "getTaskStatus":
                self.handle_get_task_status(request)
            elif request.method == "shutdown":
                self.handle_shutdown(request)
            else:
                self.send_error_response(
                    request.id, -32601, f"Method not found: {request.method}"
                )
        except Exception as e:
            self.send_error_response(
                request.id, -32603, f"Internal error: {str(e)}"
            )

    def handle_ping(self, request: JsonRpcRequest):
        """pingリクエストを処理"""
        self.send_response(request.id, {"pong": True, "timestamp": time.time()})

    def handle_execute(self, request: JsonRpcRequest):
        """RPA操作を実行"""
        params = request.params or {}

        # タスクIDを生成
        task_id = str(uuid.uuid4())
        self.current_task_id = task_id
        self.task_status[task_id] = {
            "status": "running",
            "started_at": time.time(),
            "operation": params,
        }

        # 実行開始を通知
        self.send_notification(
            "task.started",
            {"task_id": task_id, "operation": params.get("operation")},
        )

        try:
            # 非同期処理を同期的に実行
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            # 操作を実行
            category = params.get("category")
            subcategory = params.get("subcategory")
            operation = params.get("operation")
            operation_params = params.get("params", {})

            result = loop.run_until_complete(
                self.operation_manager.execute_operation(
                    category, subcategory, operation, operation_params
                )
            )

            # タスク完了
            self.task_status[task_id] = {
                "status": "completed",
                "completed_at": time.time(),
                "result": result,
            }

            # 完了を通知
            self.send_notification(
                "task.completed", {"task_id": task_id, "result": result}
            )

            # レスポンスを返す
            self.send_response(request.id, {"task_id": task_id, "result": result})

        except Exception as e:
            # エラー処理
            error_info = {
                "error": str(e),
                "traceback": traceback.format_exc(),
            }

            self.task_status[task_id] = {
                "status": "failed",
                "failed_at": time.time(),
                "error": error_info,
            }

            self.send_notification(
                "task.failed", {"task_id": task_id, "error": error_info}
            )

            self.send_error_response(
                request.id, -32000, f"Execution failed: {str(e)}"
            )

    def handle_execute_excel(self, request: JsonRpcRequest):
        """Excel操作を実行"""
        if not EXCEL_AVAILABLE:
            self.send_error_response(
                request.id, -32001, "Excel support not available"
            )
            return

        params = request.params or {}
        action = params.get("action")

        try:
            if action == "readCell":
                result = self.read_excel_cell(params)
            elif action == "writeCell":
                result = self.write_excel_cell(params)
            elif action == "readRange":
                result = self.read_excel_range(params)
            elif action == "writeRange":
                result = self.write_excel_range(params)
            else:
                self.send_error_response(
                    request.id, -32602, f"Unknown Excel action: {action}"
                )
                return

            self.send_response(request.id, result)

        except Exception as e:
            self.send_error_response(
                request.id, -32000, f"Excel operation failed: {str(e)}"
            )

    def read_excel_cell(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """Excelセルを読み込む"""
        file_path = params.get("file_path")
        sheet_name = params.get("sheet_name")
        cell = params.get("cell")

        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = (
            workbook[sheet_name] if sheet_name else workbook.active
        )
        value = sheet[cell].value
        workbook.close()

        return {"value": value}

    def write_excel_cell(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """Excelセルに書き込む"""
        file_path = params.get("file_path")
        sheet_name = params.get("sheet_name")
        cell = params.get("cell")
        value = params.get("value")

        workbook = openpyxl.load_workbook(file_path)
        sheet = (
            workbook[sheet_name] if sheet_name else workbook.active
        )
        sheet[cell] = value
        workbook.save(file_path)
        workbook.close()

        return {"success": True}

    def read_excel_range(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """Excel範囲を読み込む"""
        file_path = params.get("file_path")
        sheet_name = params.get("sheet_name")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = (
                workbook[sheet_name] if sheet_name else workbook.active
            )

            # 全データを読み込む（効率化のため最大1000行x100列に制限）
            data = []
            for row in sheet.iter_rows(max_row=1000, max_col=100):
                row_data = [cell.value for cell in row]
                # 空の行は除外
                if any(value is not None for value in row_data):
                    data.append(row_data)

            workbook.close()
            return {"data": data}

        except Exception as e:
            return {"error": str(e)}

    def write_excel_range(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """Excel範囲に書き込む"""
        file_path = params.get("file_path")
        sheet_name = params.get("sheet_name")
        start_cell = params.get("start_cell", "A1")
        data = params.get("data", [])

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = (
                workbook[sheet_name] if sheet_name else workbook.active
            )

            # データを書き込む
            from openpyxl.utils import column_index_from_string, get_column_letter

            start_col_letter = "".join(
                filter(str.isalpha, start_cell)
            )
            start_row = int("".join(filter(str.isdigit, start_cell)))
            start_col = column_index_from_string(start_col_letter)

            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    cell = sheet.cell(
                        row=start_row + row_idx,
                        column=start_col + col_idx,
                        value=value,
                    )

            workbook.save(file_path)
            workbook.close()

            return {"success": True}

        except Exception as e:
            return {"error": str(e)}

    def handle_list_operations(self, request: JsonRpcRequest):
        """利用可能な操作リストを返す"""
        operations = self.operation_manager.get_available_operations()
        self.send_response(request.id, {"operations": operations})

    def handle_get_operation_template(self, request: JsonRpcRequest):
        """操作テンプレートを返す"""
        params = request.params or {}
        category = params.get("category")
        subcategory = params.get("subcategory")
        operation = params.get("operation")

        template = self.operation_manager.get_operation_template(
            category, subcategory, operation
        )

        if template:
            self.send_response(request.id, {"template": template})
        else:
            self.send_error_response(
                request.id, -32002, f"Template not found for operation: {operation}"
            )

    def handle_cancel_task(self, request: JsonRpcRequest):
        """タスクをキャンセル"""
        params = request.params or {}
        task_id = params.get("task_id")

        if task_id in self.task_status:
            self.task_status[task_id]["status"] = "cancelled"
            self.send_response(request.id, {"cancelled": True})
        else:
            self.send_error_response(
                request.id, -32003, f"Task not found: {task_id}"
            )

    def handle_get_task_status(self, request: JsonRpcRequest):
        """タスクステータスを取得"""
        params = request.params or {}
        task_id = params.get("task_id")

        if task_id in self.task_status:
            self.send_response(request.id, self.task_status[task_id])
        else:
            self.send_error_response(
                request.id, -32003, f"Task not found: {task_id}"
            )

    def handle_shutdown(self, request: JsonRpcRequest):
        """エージェントをシャットダウン"""
        self.running = False
        self.send_response(request.id, {"shutdown": True})
        sys.exit(0)

    def send_response(self, request_id: Any, result: Any):
        """レスポンスを送信"""
        response = JsonRpcResponse(id=request_id, result=result)
        self.send_json(asdict(response))

    def send_error_response(
        self, request_id: Any, code: int, message: str, data: Any = None
    ):
        """エラーレスポンスを送信"""
        error = {"code": code, "message": message}
        if data:
            error["data"] = data
        response = JsonRpcResponse(id=request_id, error=error)
        self.send_json(asdict(response))

    def send_notification(self, method: str, params: Any = None):
        """通知を送信（レスポンス不要）"""
        notification = JsonRpcNotification(method=method, params=params)
        self.send_json(asdict(notification))

    def send_json(self, data: Dict[str, Any]):
        """JSONデータを標準出力に送信"""
        try:
            json_str = json.dumps(data, ensure_ascii=False)
            sys.stdout.write(json_str + "\n")
            sys.stdout.flush()
        except Exception as e:
            # エラーログ（通常は表示されない）
            sys.stderr.write(f"Failed to send JSON: {e}\n")
            sys.stderr.flush()


def main():
    """エントリーポイント"""
    # 標準エラー出力を無効化（JSONRPCの邪魔にならないように）
    sys.stderr = open("/dev/null", "w") if sys.platform != "win32" else open("nul", "w")

    agent = RPAAgent()
    agent.start()


if __name__ == "__main__":
    main()