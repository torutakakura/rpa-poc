"""
J_Excel カテゴリの操作
"""

from typing import Any, Dict

from .base import BaseOperation, OperationResult

# Excel操作ライブラリをオプショナルでインポート
try:
    from openpyxl import Workbook, load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelOpenOperation(BaseOperation):
    """Excelファイルを開く"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        file_path = params.get("file_path", "")
        create_if_not_exists = params.get("create_if_not_exists", False)

        error = self.validate_params(params, ["file_path"])
        if error:
            return OperationResult(status="failure", data={}, error=error)

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            import os

            file_path = os.path.expanduser(file_path)

            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                self.log(f"Opened Excel file: {file_path}")
            elif create_if_not_exists:
                wb = Workbook()
                wb.save(file_path)
                self.log(f"Created new Excel file: {file_path}")
            else:
                return OperationResult(
                    status="failure", data={}, error=f"File does not exist: {file_path}"
                )

            # ワークブックをストレージに保存
            if self.agent:
                self.agent.workbook = wb
                self.agent.workbook_path = file_path

            return OperationResult(
                status="success",
                data={
                    "file_path": file_path,
                    "sheet_count": len(wb.sheetnames),
                    "sheets": wb.sheetnames,
                },
            )
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to open Excel file: {str(e)}"
            )


class ExcelReadCellOperation(BaseOperation):
    """セルの値を読み取る"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        sheet_name = params.get("sheet_name")
        cell = params.get("cell", "A1")
        storage_key = params.get("storage_key", "")

        error = self.validate_params(params, ["cell"])
        if error:
            return OperationResult(status="failure", data={}, error=error)

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            # ワークブックを取得
            if not hasattr(self.agent, "workbook") or self.agent.workbook is None:
                return OperationResult(
                    status="failure", data={}, error="No Excel file is open"
                )

            wb = self.agent.workbook
            ws = wb[sheet_name] if sheet_name else wb.active

            # セルの値を読み取る
            value = ws[cell].value

            if storage_key:
                self.set_storage(storage_key, value)
                self.log(f"Stored cell value as '{storage_key}': {value}")

            return OperationResult(
                status="success", data={"cell": cell, "value": value, "sheet": ws.title}
            )
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to read cell: {str(e)}"
            )


class ExcelWriteCellOperation(BaseOperation):
    """セルに値を書き込む"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        sheet_name = params.get("sheet_name")
        cell = params.get("cell", "A1")
        value = params.get("value", "")

        error = self.validate_params(params, ["cell", "value"])
        if error:
            return OperationResult(status="failure", data={}, error=error)

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            # ワークブックを取得
            if not hasattr(self.agent, "workbook") or self.agent.workbook is None:
                return OperationResult(
                    status="failure", data={}, error="No Excel file is open"
                )

            wb = self.agent.workbook
            ws = wb[sheet_name] if sheet_name else wb.active

            # セルに値を書き込む
            ws[cell] = value

            self.log(f"Wrote to cell {cell}: {value}")

            return OperationResult(
                status="success", data={"cell": cell, "value": value, "sheet": ws.title}
            )
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to write cell: {str(e)}"
            )


class ExcelReadRangeOperation(BaseOperation):
    """範囲の値を読み取る"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        sheet_name = params.get("sheet_name")
        start_cell = params.get("start_cell", "A1")
        end_cell = params.get("end_cell", "C10")
        storage_key = params.get("storage_key", "")

        error = self.validate_params(params, ["start_cell", "end_cell"])
        if error:
            return OperationResult(status="failure", data={}, error=error)

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            # ワークブックを取得
            if not hasattr(self.agent, "workbook") or self.agent.workbook is None:
                return OperationResult(
                    status="failure", data={}, error="No Excel file is open"
                )

            wb = self.agent.workbook
            ws = wb[sheet_name] if sheet_name else wb.active

            # 範囲の値を読み取る
            data = []
            for row in ws[f"{start_cell}:{end_cell}"]:
                row_data = [cell.value for cell in row]
                data.append(row_data)

            if storage_key:
                self.set_storage(storage_key, data)
                self.log(f"Stored range data as '{storage_key}'")

            return OperationResult(
                status="success",
                data={
                    "range": f"{start_cell}:{end_cell}",
                    "data": data,
                    "rows": len(data),
                    "columns": len(data[0]) if data else 0,
                    "sheet": ws.title,
                },
            )
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to read range: {str(e)}"
            )


class ExcelWriteRangeOperation(BaseOperation):
    """範囲に値を書き込む"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        sheet_name = params.get("sheet_name")
        start_cell = params.get("start_cell", "A1")
        data = params.get("data", [])

        error = self.validate_params(params, ["start_cell", "data"])
        if error:
            return OperationResult(status="failure", data={}, error=error)

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            # ワークブックを取得
            if not hasattr(self.agent, "workbook") or self.agent.workbook is None:
                return OperationResult(
                    status="failure", data={}, error="No Excel file is open"
                )

            wb = self.agent.workbook
            ws = wb[sheet_name] if sheet_name else wb.active

            # 開始セルの座標を取得
            from openpyxl.utils import coordinate_to_tuple

            start_row, start_col = coordinate_to_tuple(start_cell)

            # データを書き込む
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    ws.cell(
                        row=start_row + row_idx, column=start_col + col_idx, value=value
                    )

            self.log(f"Wrote {len(data)} rows to range starting at {start_cell}")

            return OperationResult(
                status="success",
                data={
                    "start_cell": start_cell,
                    "rows_written": len(data),
                    "columns_written": len(data[0]) if data else 0,
                    "sheet": ws.title,
                },
            )
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to write range: {str(e)}"
            )


class ExcelSaveOperation(BaseOperation):
    """Excelファイルを保存"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        file_path = params.get("file_path")

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            # ワークブックを取得
            if not hasattr(self.agent, "workbook") or self.agent.workbook is None:
                return OperationResult(
                    status="failure", data={}, error="No Excel file is open"
                )

            wb = self.agent.workbook

            # 保存先を決定
            if file_path:
                import os

                file_path = os.path.expanduser(file_path)
            elif hasattr(self.agent, "workbook_path"):
                file_path = self.agent.workbook_path
            else:
                return OperationResult(
                    status="failure", data={}, error="No file path specified for saving"
                )

            # ファイルを保存
            wb.save(file_path)
            self.log(f"Saved Excel file: {file_path}")

            return OperationResult(status="success", data={"file_path": file_path})
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to save Excel file: {str(e)}"
            )


class ExcelCloseOperation(BaseOperation):
    """Excelファイルを閉じる"""

    async def execute(self, params: Dict[str, Any]) -> OperationResult:
        save_before_close = params.get("save_before_close", True)

        if not OPENPYXL_AVAILABLE:
            return OperationResult(
                status="failure",
                data={},
                error="Excel support not available. Install openpyxl.",
            )

        try:
            # ワークブックを取得
            if not hasattr(self.agent, "workbook") or self.agent.workbook is None:
                return OperationResult(
                    status="warning", data={}, error="No Excel file is open"
                )

            # 保存が必要な場合
            if save_before_close and hasattr(self.agent, "workbook_path"):
                self.agent.workbook.save(self.agent.workbook_path)
                self.log(f"Saved before closing: {self.agent.workbook_path}")

            # ワークブックを閉じる
            self.agent.workbook.close()
            self.agent.workbook = None
            self.agent.workbook_path = None

            self.log("Closed Excel file")

            return OperationResult(status="success", data={"saved": save_before_close})
        except Exception as e:
            return OperationResult(
                status="failure", data={}, error=f"Failed to close Excel file: {str(e)}"
            )
